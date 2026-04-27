import streamlit as st
import pandas as pd
import zipfile
import os
from io import BytesIO
import traceback
import re
from openpyxl import load_workbook
from collections import defaultdict

# ====== CẤU HÌNH GIAO DIỆN =======
st.set_page_config(page_title="Tạo File Hạch Toán", layout="wide")
st.title("📋 Tạo File Hạch Toán Chuẩn từ Excel")
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "🧾 Tạo File Hạch Toán", 
    "🔍 So sánh và Xoá dòng trùng", 
    "📊 File tuỳ chỉnh (Check thủ công)", 
    "📐 So sánh Số tiền giữa các file",
    "💸 Chuẩn hoá Phát sinh Nợ/Có"
])

# ====== HÀM TIỆN ÍCH CHUNG =======
def extract_month_year_from_filename(filename):
    try:
        match = re.search(r'(\d{4})[\.\-_]?\s*(\d{2})|\s*(\d{2})[\.\-_]?\s*(\d{4})', filename)
        if match:
            year = match.group(1) or match.group(4)
            month = match.group(2) or match.group(3)
            return month, year
    except: pass
    return "Tự đặt tên nhé", "Tự đặt tên nhé"

def to_ddmmyyyy(date_val):
    try:
        if pd.isnull(date_val): return ""
        if isinstance(date_val, (pd.Timestamp, pd.DatetimeTZDtype)):
            return date_val.strftime("%d/%m/%Y")
        if isinstance(date_val, (float, int)):
            return pd.to_datetime(date_val, origin='1899-12-30', unit='D').strftime("%d/%m/%Y")
        if isinstance(date_val, str):
            parsed = pd.to_datetime(date_val, dayfirst=True, errors='coerce')
            if pd.isnull(parsed): parsed = pd.to_datetime(date_val, errors='coerce')
            return parsed.strftime("%d/%m/%Y") if not pd.isnull(parsed) else ""
        return str(date_val)
    except: return ""

def format_name(name):
    clean = re.split(r'[\n\r\t\u00A0\u2003]+', str(name).strip())[0]
    clean = re.sub(r'\s+', ' ', clean)
    return clean.replace("-", "").title()

def classify_department(value, content_value=None):
    val = str(value).upper()
    if "VACCINE" in val or "VACXIN" in val: return "VACCINE"
    elif "THUỐC" in val: return "THUOC"
    elif "THẺ" in val: return "THE"
    if content_value:
        content_val = str(content_value).upper()
        if "VACCINE" in content_val: return "VACCINE"
        elif "THUỐC" in content_val: return "THUOC"
    return "KCB"

category_info = {
    "KCB": {"ma": "KHACHLE01", "ten": "Khách hàng lẻ - Khám chữa bệnh"},
    "THUOC": {"ma": "KHACHLE02", "ten": "Khách hàng lẻ - Bán thuốc"},
    "VACCINE": {"ma": "KHACHLE03", "ten": "Khách hàng lẻ - Tiêm vacxin"},
}

output_columns = [
    "Ngày hạch toán (*)", "Ngày chứng từ (*)", "Số chứng từ (*)",
    "Mã đối tượng", "Tên đối tượng", "Nộp vào TK", "Mở tại ngân hàng",
    "Lý do thu", "Diễn giải lý do thu", "Diễn giải (hạch toán)",
    "TK Nợ (*)", "TK Có (*)", "Số tiền"
]

# ====== XỬ LÝ 1 FILE =======
def process_single_file(uploaded_file, chu_hau_to, prefix):
    logs = []
    data_by_category = {cat: {} for cat in category_info}
    file_name = uploaded_file.name
    thang, nam = extract_month_year_from_filename(file_name)
    has_pos = int(nam) <= 2022 if nam.isdigit() else True

    def gen_so_chung_tu(date_str, category):
        try:
            d, m, y = date_str.split("/")
            return f"NVK{category}{d.zfill(2)}{m.zfill(2)}{y}{chu_hau_to}"
        except:
            return f"NVK_INVALID_{chu_hau_to}"

    xls = pd.ExcelFile(uploaded_file)
    for sheet_name in xls.sheet_names:
        if not sheet_name.replace(".", "", 1).isdigit() and not sheet_name.replace(",", "", 1).isdigit():
            logs.append(f"⏩ Bỏ qua sheet: {sheet_name}")
            continue
        df = xls.parse(sheet_name)
        df.columns = [str(col).strip().upper() for col in df.columns]
        if "KHOA/BỘ PHẬN" not in df.columns or "TRẢ THẺ" not in df.columns:
            logs.append(f"⚠️ Thiếu cột ở sheet {sheet_name}")
            continue
        date_column = 'NGÀY QUỸ' if 'NGÀY QUỸ' in df.columns else 'NGÀY KHÁM'
        if date_column not in df.columns:
            logs.append(f"⚠️ Không có cột ngày: {date_column}")
            continue

        df["TRẢ THẺ"] = pd.to_numeric(df["TRẢ THẺ"], errors="coerce")
        df = df[df["TRẢ THẺ"].notna() & (df["TRẢ THẺ"] != 0)]
        df = df[df[date_column].notna() & (df[date_column] != "-")]
        df = df[df["HỌ VÀ TÊN"].notna() & (df["HỌ VÀ TÊN"] != "-")]

        df["CATEGORY"] = df.apply(lambda row: classify_department(row["KHOA/BỘ PHẬN"], row.get("NỘI DUNG THU")), axis=1)

        for category in data_by_category:
            cat_df = df[df["CATEGORY"] == category]
            if cat_df.empty: continue
            for mode in ["PT", "PC"]:
                is_pt = mode == "PT"
                df_mode = cat_df[cat_df["TRẢ THẺ"] > 0] if is_pt else cat_df[cat_df["TRẢ THẺ"] < 0]
                if df_mode.empty: continue
                out_df = pd.DataFrame()
                out_df["Ngày hạch toán (*)"] = df_mode[date_column].apply(to_ddmmyyyy)
                out_df["Ngày chứng từ (*)"] = out_df["Ngày hạch toán (*)"]
                out_df["Số chứng từ (*)"] = out_df["Ngày chứng từ (*)"].apply(lambda x: gen_so_chung_tu(x, category))
                out_df["Mã đối tượng"] = category_info[category]["ma"]
                out_df["Tên đối tượng"] = df_mode["HỌ VÀ TÊN"].apply(format_name)
                out_df["Nộp vào TK"] = "1290153594"
                out_df["Mở tại ngân hàng"] = "Ngân hàng TMCP Đầu tư và Phát triển Việt Nam - Hoàng Mai"
                out_df["Lý do thu"] = ""
                ten_dv = category_info[category]['ten'].split('-')[-1].strip().lower()
                pos_phrase = " qua pos" if has_pos else ""
                out_df["Diễn giải lý do thu"] = ("Thu tiền" if is_pt else "Chi tiền") + f" {ten_dv}{pos_phrase} ngày " + out_df["Ngày chứng từ (*)"]
                out_df["TK Nợ (*)"] = "1368" if has_pos else "1121"
                out_df["Diễn giải (hạch toán)"] = out_df["Diễn giải lý do thu"] + " " + df_mode["HỌ VÀ TÊN"].apply(format_name)
                out_df["TK Có (*)"] = "131"
                out_df["Số tiền"] = df_mode["TRẢ THẺ"].abs().apply(lambda x: f"=VALUE({x})")
                out_df = out_df[output_columns]
                data_by_category[category].setdefault(sheet_name, {})[mode] = out_df
                logs.append(f"✅ {sheet_name} ({category}) [{mode}]: {len(out_df)} dòng")

    pt_pc_by_category = {cat: {"PT": defaultdict(list), "PC": defaultdict(list)} for cat in category_info}
    for category, days in data_by_category.items():
        for day, data in days.items():
            for mode in ["PT", "PC"]:
                if mode in data and not data[mode].empty:
                    pt_pc_by_category[category][mode][day].append(data[mode])

    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zip_file:
        for category in pt_pc_by_category:
            for mode in ["PT", "PC"]:
                day_dict = pt_pc_by_category[category][mode]
                if not day_dict: continue
                output = BytesIO()
                with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
                    for day, dfs in sorted(day_dict.items()):
                        merged_df = pd.concat(dfs).reset_index(drop=True)
                        if merged_df.empty: continue
                        merged_df.to_excel(writer, sheet_name=day.strip(), index=False)
                        workbook = writer.book
                        worksheet = writer.sheets[day.strip()]
                        header_format = workbook.add_format({'bold': True, 'bg_color': '#D9E1F2', 'border': 1})
                        for col_num, col_name in enumerate(merged_df.columns):
                            worksheet.write(0, col_num, col_name, header_format)
                        for i, col in enumerate(merged_df.columns):
                            max_width = max([len(str(col))] + [len(str(v)) for v in merged_df[col].values])
                            worksheet.set_column(i, i, max_width + 2)
                        worksheet.set_tab_color('#92D050')
                output.seek(0)
                file_path = f"{prefix}_{category}/{mode}.xlsx"
                zip_file.writestr(file_path, output.read())

    cleaned_zip = BytesIO()
    with zipfile.ZipFile(zip_buffer, "r") as zin, zipfile.ZipFile(cleaned_zip, "w") as zout:
        for item in zin.infolist():
            if item.filename.endswith(".xlsx"):
                with zin.open(item.filename) as f:
                    wb = load_workbook(f, data_only=False)
                    for sheet in wb.worksheets:
                        headers = [cell.value for cell in sheet[1]]
                        if "Số tiền" in headers:
                            col_idx = headers.index("Số tiền") + 1
                            for row in sheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                                cell = row[0]
                                if cell.data_type == "f" and isinstance(cell.value, str):
                                    match = re.search(r"=VALUE\(([\d.]+)\)", cell.value)
                                    if match:
                                        cell.value = float(match.group(1))
                                        cell.data_type = 'n'
                    temp_output = BytesIO()
                    wb.save(temp_output)
                    temp_output.seek(0)
                    zout.writestr(item.filename, temp_output.read())
            else:
                zout.writestr(item, zin.read(item.filename))

    return cleaned_zip, logs


# ====== GIAO DIỆN TAB 1 =======
with tab1:
    uploaded_files = st.file_uploader("📂 Chọn nhiều file Excel (.xlsx)", type=["xlsx"], accept_multiple_files=True)
    chu_hau_to = st.text_input("✍️ Hậu tố chứng từ (VD: A, B1, NV123)").strip().upper()

    if st.button("🚀 Tạo File Zip Tổng Hợp") and uploaded_files and chu_hau_to:
        try:
            zip_master = BytesIO()
            logs_all = []
            with zipfile.ZipFile(zip_master, "w") as zip_all:
                for uploaded_file in uploaded_files:
                    file_name = uploaded_file.name
                    thang, nam = extract_month_year_from_filename(file_name)
                    prefix = f"T{thang}_{nam}" if thang != "Tự đặt tên nhé" and nam != "Tự đặt tên nhé" else "TBD"
                    zip_sub, logs = process_single_file(uploaded_file, chu_hau_to, prefix)
                    folder_name = f"{os.path.splitext(file_name)[0]}_{prefix}"
                    with zipfile.ZipFile(zip_sub, "r") as zsub:
                        for item in zsub.infolist():
                            content = zsub.read(item.filename)
                            zip_all.writestr(f"{folder_name}/{item.filename}", content)
                    logs_all.append(f"📄 {file_name}:")
                    logs_all.extend([f"  - {line}" for line in logs])

            st.success("🎉 Tạo file zip tổng hợp thành công!")
            st.download_button("📦 Tải File Zip Tổng", data=zip_master.getvalue(), file_name=f"TongHop_{chu_hau_to}.zip")
            st.markdown("### 📑 Nhật ký xử lý toàn bộ:")
            st.text("\n".join(logs_all))
        except Exception as e:
            st.error("❌ Đã xảy ra lỗi:")
            st.code(traceback.format_exc(), language="python")

with tab2:
    st.header("🔍 So sánh với File Gốc và Xoá dòng trùng")

    base_file = st.file_uploader("📂 File Gốc (Base - Excel)", type=["xlsx"], key="base_file")
    zip_compare_file = st.file_uploader("📦 File ZIP đầu ra của hệ thống", type=["zip"], key="zip_compare")
    chu_hau_to = st.text_input("✍️ Hậu tố chứng từ", value="DA").strip().upper()

    # === Helper hàm chuẩn hóa ===
    def normalize_name(name):
        return re.sub(r'\s+', ' ', str(name).strip().lower())

    def normalize_date(date_val):
        try:
            d = pd.to_datetime(date_val, dayfirst=True, errors="coerce")
            return d.strftime("%d/%m/%Y") if pd.notnull(d) else None
        except:
            return None

    def normalize_columns(columns):
        return [
            str(c).strip()
            .replace('\xa0', ' ')
            .replace('\n', ' ')
            .replace('\t', ' ')
            .replace('\r', ' ')
            .strip()
            for c in columns
        ]

    def extract_type_from_path(path):
        path = path.upper()
        if "KCB" in path: return "KCB"
        elif "THUOC" in path: return "THUOC"
        elif "VACCINE" in path: return "VACCINE"
        elif "THE" in path: return "THE"
        return "KHAC"

    def format_sct(row, chu_hau_to="DA"):
        try:
            d = pd.to_datetime(row["Ngày chứng từ (*)"], dayfirst=True)
            date_str = d.strftime("%d%m%y")
            month = d.strftime("%m")
            is_pt = "THU" in str(row.get("Diễn giải", "")).upper()
            mode = "PT" if is_pt else "PC"
            return f"NVK/{mode}{month}_{date_str}_{chu_hau_to}"
        except:
            return f"NVK/PT??_TBD_{chu_hau_to}"

    # === Nút xử lý chính ===
    if st.button("🚫 Xoá dòng trùng theo Tên + Ngày + Số Tiền"):
        if base_file and zip_compare_file:
            try:
                base_df = pd.read_excel(base_file)
                base_df.columns = normalize_columns(base_df.columns)

                # Chuẩn hóa base
                base_df["Tên chuẩn"] = base_df["Tên đối tượng"].apply(normalize_name)
                base_df["Ngày chuẩn"] = base_df["Ngày hạch toán (*)"].apply(normalize_date)
                base_df["Số tiền chuẩn"] = pd.to_numeric(base_df["Số tiền"], errors='coerce')
                base_keys = set(zip(base_df["Tên chuẩn"], base_df["Ngày chuẩn"], base_df["Số tiền chuẩn"]))

                # Mở file ZIP
                zin = zipfile.ZipFile(zip_compare_file, 'r')
                output_zip = BytesIO()

                logs = []
                total_removed = 0

                with zipfile.ZipFile(output_zip, "w") as zout:
                    for file_name in zin.namelist():
                        if not file_name.endswith(".xlsx"):
                            continue

                        with zin.open(file_name) as f:
                            xls = pd.ExcelFile(f)
                            output = BytesIO()

                            with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
                                for sheet in xls.sheet_names:
                                    df = pd.read_excel(xls, sheet_name=sheet)
                                    df.columns = normalize_columns(df.columns)

                                    if not set(["Tên đối tượng", "Ngày hạch toán (*)", "Số tiền"]).issubset(df.columns):
                                        continue

                                    df["Tên chuẩn"] = df["Tên đối tượng"].apply(normalize_name)
                                    df["Ngày chuẩn"] = df["Ngày hạch toán (*)"].apply(normalize_date)
                                    df["Số tiền chuẩn"] = pd.to_numeric(df["Số tiền"], errors='coerce')

                                    before = len(df)
                                    df = df[~df.set_index(["Tên chuẩn", "Ngày chuẩn", "Số tiền chuẩn"]).index.isin(base_keys)]
                                    removed = before - len(df)
                                    total_removed += removed
                                    logs.append(f"- {file_name} | {sheet}: ❌ Đã xoá {removed} dòng")

                                    # Đổi cột Diễn giải lý do thu → Diễn giải
                                    if "Diễn giải lý do thu" in df.columns:
                                        df.rename(columns={"Diễn giải lý do thu": "Diễn giải"}, inplace=True)

                                    # Gán số chứng từ mới
                                    df["Số chứng từ (*)"] = df.apply(lambda row: format_sct(row, chu_hau_to), axis=1)

                                    # Chỉ giữ cột cần
                                    keep_cols = [
                                        "Ngày hạch toán (*)", "Ngày chứng từ (*)", "Số chứng từ (*)",
                                        "Mã đối tượng có", "Mã đối tượng nợ", "Diễn giải",
                                        "Tên đối tượng", "Diễn giải (hạch toán)", "TK Nợ (*)",
                                        "TK Có (*)", "Số tiền"
                                    ]
                                    df = df[[c for c in keep_cols if c in df.columns]]

                                    # Sửa tên cột "Tên đối tượng" → "họ tên"
                                    df.rename(columns={"Tên đối tượng": "họ tên"}, inplace=True)

                                    df.to_excel(writer, sheet_name=sheet, index=False)
                                    worksheet = writer.sheets[sheet]
                                    for i, col in enumerate(df.columns):
                                        max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
                                        worksheet.set_column(i, i, max_len)

                            output.seek(0)
                            mode = "PT" if "PT" in file_name.upper() else "PC"
                            loai = extract_type_from_path(file_name)
                            zout.writestr(f"{mode}_{loai}.xlsx", output.read())

                # Hiển thị kết quả
                st.success(f"✅ Đã xoá {total_removed} dòng trùng khớp từ ZIP.")
                st.download_button("📥 Tải ZIP sau khi xoá", data=output_zip.getvalue(), file_name="cleaned_output.zip")

                st.markdown("### 📜 Nhật ký xử lý:")
                for log in logs:
                    st.markdown(log)

            except Exception as e:
                st.error("❌ Lỗi khi xử lý:")
                st.code(traceback.format_exc(), language="python")

with tab3:
    st.header("📊 Gộp Dữ Liệu Tháng Thành 1 File Excel Tổng Hợp")
    zip_input = st.file_uploader("📂 Tải lên file Zip đầu ra từ Tab 1", type=["zip"], key="zip_monthly")

    if zip_input:
        try:
            group_data = {
                "PT_KCB": [], "PC_KCB": [],
                "PT_THUOC": [], "PC_THUOC": [],
                "PT_VACCINE": [], "PC_VACCINE": []
            }

            # 🧠 Lấy tên tháng & năm từ file zip nếu có thể
            match = re.search(r't(\d{1,2})[_\-\.](\d{4})', zip_input.name.lower())
            if match:
                thang_text = f"T{int(match.group(1))}"
                nam_text = match.group(2)
            else:
                thang_text = "TBD"
                nam_text = "XXXX"

            with zipfile.ZipFile(zip_input, "r") as zipf:
                for filename in zipf.namelist():
                    if not filename.endswith(".xlsx"):
                        continue

                    with zipf.open(filename) as f:
                        xls = pd.ExcelFile(f)
                        for sheet_name in xls.sheet_names:
                            if sheet_name.startswith("PT") or sheet_name.startswith("PC"):
                                df = xls.parse(sheet_name)
                                if not set(["Ngày chứng từ (*)", "Tên đối tượng", "Số tiền"]).issubset(df.columns):
                                    continue

                                short_type = None
                                if "KCB" in filename.upper():
                                    short_type = "KCB"
                                elif "THUOC" in filename.upper():
                                    short_type = "THUOC"
                                elif "VACCINE" in filename.upper():
                                    short_type = "VACCINE"
                                else:
                                    continue

                                mode = "PT" if sheet_name.startswith("PT") else "PC"
                                key = f"{mode}_{short_type}"

                                df_filtered = df[["Ngày chứng từ (*)", "Tên đối tượng", "Số tiền"]].copy()
                                group_data[key].append(df_filtered)

            output = BytesIO()
            with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
                for key, df_list in group_data.items():
                    if not df_list:
                        continue
                    merged_df = pd.concat(df_list, ignore_index=True)
                    merged_df.columns = ["Ngày", "Tên", "Số tiền"]

                    # Thêm công thức cột Ghi chú
                    merged_df["Ghi chú"] = ""

                    merged_df.to_excel(writer, sheet_name=key, index=False, startrow=0, header=True)

                    workbook = writer.book
                    worksheet = writer.sheets[key]

                    # Format header
                    header_format = workbook.add_format({'bold': True, 'bg_color': '#FCE4D6', 'border': 1})
                    for col_num, value in enumerate(merged_df.columns):
                        worksheet.write(0, col_num, value, header_format)
                        max_width = max(len(str(value)), *(merged_df.iloc[:, col_num].astype(str).map(len)))
                        worksheet.set_column(col_num, col_num, max_width + 2)

                    # Viết công thức Ghi chú
                    for row_num in range(1, len(merged_df)+1):
                        formula = f'=IF(COUNTIFS(A:A,A{row_num+1},B:B,B{row_num+1},C:C,C{row_num+1})>1,"Lặp","")'
                        worksheet.write_formula(row_num, 3, formula)

                    worksheet.set_tab_color("#FFD966")

            file_name_out = f"TongHop_{thang_text}_{nam_text}.xlsx"
            st.success(f"🎉 Đã gộp xong dữ liệu tháng {thang_text}/{nam_text}!")
            st.download_button("📥 Tải File Tổng Hợp", data=output.getvalue(), file_name=file_name_out)

        except Exception as e:
            st.error("❌ Lỗi khi xử lý file Zip:")
            st.code(traceback.format_exc(), language="python")

with tab4:
    st.subheader("📑 So sánh 'Số tiền' giữa nhiều file Excel")

    uploaded_excels = st.file_uploader(
        "📂 Chọn nhiều file Excel để so sánh", 
        type=["xlsx"], 
        accept_multiple_files=True, 
        key="multi_excel_compare"
    )

    if uploaded_excels:
        try:
            all_records = []

            for file in uploaded_excels:
                xl = pd.ExcelFile(file)
                for sheet in xl.sheet_names:
                    df = xl.parse(sheet)
                    df.columns = [str(c).strip() for c in df.columns]

                    cols_lower = [c.lower() for c in df.columns]
                    required = {"số tiền", "số chứng từ (*)", "ngày chứng từ (*)"}
                    if not required.issubset(set(cols_lower)):
                        continue

                    ten_col = next((c for c in df.columns if c.strip().lower() in ["họ và tên", "tên đối tượng"]), None)
                    if not ten_col: continue

                    df["TÊN FILE"] = file.name
                    df["TÊN SHEET"] = sheet
                    df["KEY"] = df[ten_col].astype(str).str.strip() + "_" + df["Số chứng từ (*)"].astype(str)

                    df["SỐ TIỀN GỐC"] = (
                        df["Số tiền"]
                        .astype(str)
                        .str.replace("=VALUE(", "", regex=False)
                        .str.replace(")", "", regex=False)
                        .astype(float)
                    )

                    all_records.append(df[["KEY", "SỐ TIỀN GỐC", "TÊN FILE", "TÊN SHEET"]])

            if not all_records:
                st.warning("⚠️ Không tìm thấy dữ liệu phù hợp để so sánh.")
            else:
                full_df = pd.concat(all_records)
                pivot_df = full_df.pivot_table(
                    index="KEY", 
                    columns="TÊN FILE", 
                    values="SỐ TIỀN GỐC", 
                    aggfunc="first"
                ).reset_index()

                # So sánh: những dòng có sự khác biệt giữa các file
                diff_mask = pivot_df.drop("KEY", axis=1).apply(
                    lambda row: len(set(row.dropna())) > 1, axis=1
                )
                result_df = pivot_df[diff_mask]

                st.markdown(f"""
                ### 📊 Kết quả so sánh 'Số tiền'
                - Tổng dòng dữ liệu: `{len(pivot_df)}`
                - Số dòng khác biệt: `{len(result_df)}`
                """)

                st.dataframe(result_df, use_container_width=True)

                excel_bytes = BytesIO()
                with pd.ExcelWriter(excel_bytes, engine="xlsxwriter") as writer:
                    result_df.to_excel(writer, index=False)
                excel_bytes.seek(0)

                st.download_button(
                    "⬇️ Tải kết quả so sánh (Excel)",
                    data=excel_bytes.getvalue(),
                    file_name="So_sanh_So_tien.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        except Exception as e:
            st.error("❌ Đã xảy ra lỗi khi xử lý các file Excel:")
            st.code(traceback.format_exc(), language="python")

with tab5:
    st.subheader("📌 So khớp công nợ giữa MISA và Excel Thu thực tế")

    col1, col2 = st.columns(2)
    with col1:
        misa_input = st.text_area("📥 Dán bảng từ MISA", height=300, help="Copy toàn bộ bảng công nợ từ phần mềm MISA (bao gồm cả cột Phát sinh, Số dư, Nợ/Có...)")
    with col2:
        excel_input = st.text_area("📥 Dán bảng từ Excel", height=300, help="Copy bảng Excel thu tiền (bao gồm Họ tên, Mã Y tế, Số tiền, Tiền mặt, Trả thẻ...)")

    if misa_input and excel_input:
        try:
            def parse_misa(text):
                lines = [line.strip() for line in text.splitlines() if line.strip()]
                records = []
                current_ma = ""
                for line in lines:
                    # Dòng chứa mã y tế
                    if re.match(r'^[A-Z0-9]{6,}', line):
                        current_ma = line.strip()
                    elif re.match(r'^\d{2}/\d{2}/\d{4}', line):
                        parts = line.split("\t")
                        if len(parts) >= 10:
                            so_phat_sinh = parts[10].replace('.', '').replace(',', '.').strip()
                            try:
                                ps = float(so_phat_sinh)
                                records.append({
                                    'Mã Y Tế': current_ma,
                                    'Phát sinh': ps
                                })
                            except:
                                pass
                return pd.DataFrame(records)

            def parse_excel(text):
                lines = [line.strip() for line in text.splitlines() if line.strip()]
                records = []
                for line in lines:
                    parts = line.split("\t")
                    if len(parts) >= 6:
                        try:
                            ho_ten = parts[4].strip().upper()
                            ma_y_te = parts[5].strip().replace('.', '')
                            so_tien = parts[6].replace(',', '').strip()
                            thu_thuc_te = float(so_tien)
                            records.append({
                                'Mã Y Tế': ma_y_te,
                                'Họ tên': ho_ten,
                                'Đã thu': thu_thuc_te
                            })
                        except:
                            continue
                return pd.DataFrame(records)

            df_misa = parse_misa(misa_input)
            df_excel = parse_excel(excel_input)

            df = pd.merge(df_misa, df_excel, on="Mã Y Tế", how="outer")
            df["Phát sinh"] = df["Phát sinh"].fillna(0)
            df["Đã thu"] = df["Đã thu"].fillna(0)
            df["Chênh lệch"] = df["Đã thu"] - df["Phát sinh"]

            st.success("✅ Đã so khớp xong.")
            st.dataframe(df)

            output = BytesIO()
            df.to_excel(output, index=False)
            st.download_button("⬇️ Tải kết quả so khớp", data=output.getvalue(), file_name="so_khop_cong_no.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        except Exception as e:
            st.error("❌ Lỗi xử lý dữ liệu:")
            st.exception(e)
